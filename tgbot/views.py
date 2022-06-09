import json
import logging
from django.views import View
from django.views.generic.detail import DetailView
from django.views.generic import TemplateView
from django.views.generic.list import ListView
from django.views.generic.edit import UpdateView, CreateView
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib.auth import authenticate, login
from django.contrib.auth.views import LogoutView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.core import serializers

from django.template import RequestContext

from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.utils.html import strip_tags

from dtb.settings import DEBUG, LOGIN_URL
from tgbot.dispatcher import process_telegram_event, TELEGRAM_BOT_USERNAME

from .forms import CustomAuthForm, VacancyForm, EmployeeValuesForm
from .models import Profile, User, Experience, Vacancy, EmployeeValues, ProfileStatuses, ProfileStatusHistory

logger = logging.getLogger(__name__)


def index(request):
    return redirect('signin')


class LogoutView(LogoutView):
    next_page = 'tgbot:signin'


class TelegramBotWebhookView(View):
    # WARNING: if fail - Telegram webhook will be delivered again.
    # Can be fixed with async celery task execution
    def post(self, request, *args, **kwargs):
        if DEBUG:
            process_telegram_event(json.loads(request.body))
        else:
            # Process Telegram event in Celery worker (async)
            # Don't forget to run it and & Redis (message broker for Celery)!
            # Read Procfile for details
            # You can run all of these services via docker-compose.yml
            process_telegram_event.delay(json.loads(request.body))

        # TODO: there is a great trick to send action in webhook response
        # e.g. remove buttons, typing event
        return JsonResponse({"ok": "POST request processed"})

    def get(self, request, *args, **kwargs):  # for debug
        return JsonResponse({"ok": "Get request received! But nothing done"})


class Cabinet(LoginRequiredMixin, TemplateView):
    template_name = 'tgbot/cabinet.html'
    # login_url = 'signin'

    def get(self, request):
        return render(request, 'tgbot/cabinet.html', {'context': self.get_context_data()})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['profiles'] = Profile.objects.all()
        context['vacancies'] = Vacancy.objects.filter(is_active=True).order_by('id')
        return context


class SignInView(View):
    def get(self, request):
        return render(request, 'tgbot/signin.html', {'form':CustomAuthForm()})

    def post(self, request):
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'tgbot/signin.html', {'form':CustomAuthForm(), 'error':'Account didn\'t found!'})
        else:
            login(request, user)
            return redirect('tgbot:cabinet')


class ProfilelListView(LoginRequiredMixin, ListView):
    template_name = 'tgbot/profile_list.html'
    queryset = Profile.objects.all()
    context_object_name = 'profiles'
    model = Profile
    paginate_by = 12

    def get_queryset(self):
        filter_val1 = self.request.GET.get('status', 'all')
        filter_val2 = self.request.GET.get('vacancy', '-1')
        exps = Experience.objects.filter(vacancy_id=int(filter_val2)) if filter_val2 != '-1' else Experience.objects.all()
        exps = [exp.user.user_id for exp in exps]
        order = self.request.GET.get('orderby', 'salary_await')
        if filter_val1 == 'all' and filter_val2 == '-1':
            queryset = Profile.objects.all()
            return queryset.order_by(order)
        queryset = Profile.objects.filter(
            status=filter_val1,
            user__user_id__in=exps
        ) if filter_val1 != 'all' else Profile.objects.filter(
            user__user_id__in=exps
        )
        return queryset.order_by(order)

    def get_context_data(self, **kwargs):
        context = super(ProfilelListView, self).get_context_data(**kwargs)
        context['status'] = self.request.GET.get('status', 'all')
        context['vacancy'] = self.request.GET.get('vacancy', 'all')
        context['orderby'] = self.request.GET.get('orderby', 'pk')
        st_tuple = (('all', 'Все'),)
        statuses = Profile.STATUSES + st_tuple
        context['statuses'] = statuses
        vacancies = [{'id':str(it.id), 'title':it.title} for it in Vacancy.objects.all()]
        vacancies.append({'id':'-1', 'title':'Все'})
        context['vacancies'] = vacancies
        return context


class ProfileDetailView(LoginRequiredMixin, DetailView):
    template_name = 'tgbot/profile_detail.html'
    context_object_name = 'profile'
    model = Profile

    def get_context_data(self, **kwargs):
        context = super(ProfileDetailView, self).get_context_data(**kwargs)
        context['statuses'] = self.object.get_statuses_history_set()
        return context


class ProfileActionView(LoginRequiredMixin, View):
    #model = Profile

    def post(self, request, pk):
        status = self.request.POST.get('status', '')
        statuses = [el.id for el in ProfileStatuses.objects.all()]
        if int(status) in statuses:
            profile = self.get_object(pk)
            current_stage = profile.get_current_stage()
            profile.change_status(status, current_stage, request.user)
        return redirect(profile)

    def get_object(self, pk):
        obj = get_object_or_404(Profile, pk=pk)
        return obj


@login_required
def export_profiles_to_xlsx(request):
    prof_ids = request.GET.getlist('profs', '')
    prof_ids = [int(prof) for prof in prof_ids]
    """
    Downloads filtred Profiles as Excel file with a single worksheet
    """
    profile_queryset = Profile.objects.filter(user__user_id__in=prof_ids)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-profiles.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Candidates'

    # Define the titles for columns
    columns = [
        'UserTG',
        'Name Family',
        'Phone',
        'Working',
        'Salary await',
        'Status',
        'Values',
        'Experience',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for profile in profile_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            profile.pk,
            profile.name_family,
            profile.phone,
            profile.working,
            profile.salary_await,
            profile.get_status_display(),
            profile.emp_values,
            strip_tags(profile.exp)
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


class VacancyListView(LoginRequiredMixin, ListView):
    template_name = 'tgbot/vacancy_list.html'
    queryset = Vacancy.objects.all()
    context_object_name = 'vacancies'
    model = Vacancy
    paginate_by = 12


class VacancyDetailView(LoginRequiredMixin, DetailView):
    template_name = 'tgbot/vacancy_detail.html'
    context_object_name = 'vacancy'
    model = Vacancy


class VacancyUpdateView(LoginRequiredMixin, UpdateView):
    model = Vacancy
    form_class = VacancyForm
    template_name = 'tgbot/vacancy_update.html'


class VacancyCreateView(LoginRequiredMixin, CreateView):
    model = Vacancy
    form_class = VacancyForm
    template_name = 'tgbot/vacancy_create.html'


class EmployeeValuesListView(LoginRequiredMixin, ListView):
    template_name = 'tgbot/employee_values_list.html'
    queryset = EmployeeValues.objects.filter(is_base=True)
    context_object_name = 'emp_values'
    model = EmployeeValues
    paginate_by = 12


class EmployeeValuesCreateView(LoginRequiredMixin, CreateView):
    model = EmployeeValues
    form_class = EmployeeValuesForm
    template_name = 'tgbot/employee_values_create.html'

    def get_initial(self):
        initial = super(EmployeeValuesCreateView, self).get_initial()
        initial = initial.copy()
        initial['is_base'] = True
        return initial


class EmployeeValuesUpdateView(LoginRequiredMixin, UpdateView):
    model = EmployeeValues
    form_class = EmployeeValuesForm
    template_name = 'tgbot/employee_values_update.html'


class StageCommentAdd(LoginRequiredMixin, View):
    def post(self, request, pk):
        stage_id = self.request.POST.get('stage_id', '')
        comment = self.request.POST.get('stage_comment', '')
        stage = ProfileStatusHistory.objects.get(pk=int(stage_id))
        statuses = [el.id for el in ProfileStatuses.objects.all()]
        stage.comment = comment
        stage.save()
        return redirect(stage.profile)

    def get_object(self, pk):
        obj = get_object_or_404(ProfileStatusHistory, pk=pk)
        return obj


class ProfileCommentAdd(LoginRequiredMixin, View):
    def post(self, request, pk):
        profile = self.get_object(pk)
        comment = self.request.POST.get('profile_comment', '')
        profile.hr_comment = comment
        profile.save()
        return redirect(profile)

    def get_object(self, pk):
        obj = get_object_or_404(Profile, pk=pk)
        return obj


def error_404(request, *args, **argv):
    return render(request, 'tgbot/404.html', status=404)


def error_500(request, *args, **argv):
    return render(request, 'tgbot/500.html', status=500)